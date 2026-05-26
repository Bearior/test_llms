"""Document label catalog.

Loads the 116 business labels from `thai_documents_list.xlsx`, groups them into
visual families, and collapses duplicate labels into a canonical set of
``physical_type`` classes (the classifier's output space).

Key design decision (see plan ADR):
  * The xlsx has 116 rows but only ~94 distinct names — the duplicates are the
    same physical document reused for different *business purposes*.
  * The classifier predicts ``physical_type`` (the canonical, de-duplicated
    class). The original 116 business labels are recovered at the application
    layer by pairing ``physical_type`` with a ``business_role`` (set from
    workflow context, NOT inferred from the image).
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from functools import lru_cache
from typing import Optional

from openpyxl import load_workbook

from .config import XLSX_PATH


# --- Visual families ---------------------------------------------------------
# A document's family is what a coarse visual classifier can plausibly tell
# apart; the exact subtype within a family usually requires *reading text*.
FAMILY_PASSPORT = "PASSPORT_BOOKLET"
FAMILY_VISA = "VISA_STAMP"
FAMILY_ID_CARD = "ID_CARD"
FAMILY_HOUSEHOLD = "HOUSEHOLD_REGISTRATION"
FAMILY_PERMIT = "PERMIT_LICENSE"
FAMILY_CERT_LETTER = "CERTIFICATE_LETTER"
FAMILY_OTHER = "OTHER"

ALL_FAMILIES = [
    FAMILY_PASSPORT,
    FAMILY_VISA,
    FAMILY_ID_CARD,
    FAMILY_HOUSEHOLD,
    FAMILY_PERMIT,
    FAMILY_CERT_LETTER,
    FAMILY_OTHER,
]

# Sentinel for the abstention / out-of-distribution route.
UNKNOWN_TYPE = "__UNKNOWN__"


@dataclass(frozen=True)
class DocumentType:
    """A canonical physical document type (classifier output)."""
    physical_type: str           # canonical normalized name (stable key)
    display_name: str            # original human-readable name
    family: str
    subtype_marker: Optional[str] = None  # short code that distinguishes subtypes
    source_rows: tuple = field(default_factory=tuple)  # xlsx "No" values collapsed here


def _normalize(name: str) -> str:
    """Canonicalize a label for de-duplication: NFC, collapse whitespace, strip."""
    n = unicodedata.normalize("NFC", str(name)).strip()
    n = re.sub(r"\s+", " ", n)
    return n


# --- Family detection (keyword heuristics on the label text) -----------------
def detect_family(name: str) -> str:
    n = name.lower()
    th = name  # Thai matching keeps original

    # Visa stamps / labels: explicit VISA, Smart "X", LTR "X", Diplomatic/Official
    if ("visa" in n or "smart " in n or n.startswith("ltr ")
            or "diplomatic" in n or "official " in n):
        return FAMILY_VISA
    # Passport-style travel booklets
    if ("passport" in n or "travel document" in n
            or "certificate of identity" in n or "หนังสือเดินทาง" in th):
        return FAMILY_PASSPORT
    # ID-style cards
    if ("id card" in n or "บัตรประจำตัว" in th or "บัตรประชาชน" in th
            or "สมุดประจำตัวคนต่างด้าว" in th):
        return FAMILY_ID_CARD
    # Household registration forms (ทร.13/14, ทร.13/1, ทร.14/1)
    if ("ทะเบียนบ้าน" in th or "ทะเบียนราษฎร" in th or "ทร.1" in th
            or "ทะเบียนประวัติ" in th):
        return FAMILY_HOUSEHOLD
    # Permits / licenses
    if ("work permit" in n or "ใบอนุญาต" in th or "driver" in n
            or "border pass" in n or "หนังสือผ่านแดน" in th
            or "หนังสือคนประจำเรือ" in th):
        return FAMILY_PERMIT
    # Certificates / letters / supporting docs
    if ("certificate" in n or "หนังสือรับรอง" in th or "หนังสือ" in th
            or "สัญญา" in th or "ใบเสร็จ" in th or "สลิป" in th
            or "แบบยื่น" in th or "แบบรับรอง" in th or "คำสั่งศาล" in th
            or "เอกสาร" in th or "บัตร" in th or "title deed" in n
            or "กรณี" in th):
        return FAMILY_CERT_LETTER
    return FAMILY_OTHER


# --- Subtype marker extraction ----------------------------------------------
_VISA_CODE_RE = re.compile(r'"([^"]+)"')


def extract_subtype_marker(name: str) -> Optional[str]:
    """Pull the short distinguishing code out of a label, when present.

    e.g. Non-Immigrant VISA "B" -> 'B' ; Smart "T" -> 'T' ; LTR "W" -> 'W'.
    These quoted codes are exactly the near-identical-subtype problem.
    """
    m = _VISA_CODE_RE.search(name)
    if m:
        return m.group(1).strip()
    # Household registration variants encode the marker in the form id.
    m2 = re.search(r"(ทร\.\s*\d+(?:/\d+)?)", name)
    if m2:
        return m2.group(1).replace(" ", "")
    return None


@lru_cache(maxsize=1)
def load_raw_labels() -> list[tuple[int, str]]:
    """Return the raw [(No, Document Name), ...] rows from the xlsx."""
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows: list[tuple[int, str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if row is None or row[1] is None:
            continue
        no = row[0]
        name = _normalize(row[1])
        if name:
            rows.append((int(no) if no is not None else -1, name))
    wb.close()
    return rows


@lru_cache(maxsize=1)
def build_catalog() -> list[DocumentType]:
    """Build the canonical de-duplicated DocumentType catalog (physical types)."""
    raw = load_raw_labels()
    by_name: dict[str, dict] = {}
    for no, name in raw:
        rec = by_name.setdefault(name, {"rows": [], "name": name})
        rec["rows"].append(no)

    catalog: list[DocumentType] = []
    for name, rec in by_name.items():
        catalog.append(
            DocumentType(
                physical_type=name,            # normalized name is the stable key
                display_name=name,
                family=detect_family(name),
                subtype_marker=extract_subtype_marker(name),
                source_rows=tuple(rec["rows"]),
            )
        )
    # Stable, deterministic ordering.
    catalog.sort(key=lambda d: d.physical_type)
    return catalog


def canonical_count() -> int:
    return len(build_catalog())


def business_label_count() -> int:
    return len(load_raw_labels())


def alias_map() -> dict[str, str]:
    """Map every original (normalized) business label -> canonical physical_type.

    Because we normalize on the same key, this is the identity for distinct
    names and collapses exact duplicates. Extend this table when business units
    use genuinely different *strings* for the same physical document.
    """
    return {dt.display_name: dt.physical_type for dt in build_catalog()}


def types_by_family(family: str) -> list[DocumentType]:
    return [dt for dt in build_catalog() if dt.family == family]


def get_type(physical_type: str) -> Optional[DocumentType]:
    for dt in build_catalog():
        if dt.physical_type == physical_type:
            return dt
    return None
